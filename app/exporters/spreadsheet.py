"""XLSX spreadsheet generator using openpyxl."""

import os
import tempfile
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Shared styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

EXPORT_DIR = os.environ.get("EXPORT_DIR", os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "exports"))


def _apply_header_style(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _auto_width(ws, max_width: int = 50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:100]:  # sample first 100 rows
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _write_rows(ws, headers: list[str], rows: list[list[Any]]):
    ws.append(headers)
    _apply_header_style(ws, len(headers))
    for row in rows:
        ws.append(row)


def generate_sitemap_xlsx(results: list[dict], domain: str, summary: dict) -> str:
    """Generate sitemap export XLSX. Returns file path."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    wb = Workbook()

    # Sheet 1: URLs
    ws_urls = wb.active
    ws_urls.title = "URLs"
    url_headers = [
        "URL", "Status Code", "Content-Type", "Page Title", "Title Length",
        "Meta Description", "Description Length", "H1 Text", "Canonical URL",
        "Canonical Match", "Robots Meta", "Is Indexable", "Robots.txt Status",
        "Lastmod", "Changefreq", "Priority", "Word Count",
        "OG Title", "OG Description", "Extraction Method", "Notes", "Error",
    ]
    ws_urls.append(url_headers)
    _apply_header_style(ws_urls, len(url_headers))

    for r in results:
        row = [
            r.get("url", ""),
            r.get("status_code", ""),
            r.get("content_type", ""),
            r.get("title", ""),
            r.get("title_length", ""),
            r.get("meta_description", ""),
            r.get("description_length", ""),
            r.get("h1", ""),
            r.get("canonical", ""),
            r.get("canonical_match", ""),
            r.get("robots_meta", ""),
            r.get("is_indexable", ""),
            r.get("robots_txt_status", "Allowed"),
            r.get("lastmod", ""),
            r.get("changefreq", ""),
            r.get("priority", ""),
            r.get("word_count", ""),
            r.get("og_title", ""),
            r.get("og_description", ""),
            r.get("extraction_method", ""),
            r.get("notes", ""),
            r.get("error", ""),
        ]
        ws_urls.append(row)

    # Apply conditional formatting for errors/warnings
    for row_idx in range(2, ws_urls.max_row + 1):
        status_cell = ws_urls.cell(row=row_idx, column=2)
        if status_cell.value and isinstance(status_cell.value, int):
            if status_cell.value >= 400:
                status_cell.fill = ERROR_FILL
            elif status_cell.value >= 300:
                status_cell.fill = WARNING_FILL
            elif status_cell.value == 200:
                status_cell.fill = PASS_FILL

    _auto_width(ws_urls)

    # Sheet 2: Summary
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ["Domain", domain],
        ["Export Date", datetime.now(timezone.utc).isoformat()],
        ["Total URLs", summary.get("total_urls", 0)],
        ["URLs Fetched Successfully", summary.get("fetched_ok", 0)],
        ["URLs Failed", summary.get("failed", 0)],
        ["HTTP 200", summary.get("status_200", 0)],
        ["HTTP 301/302", summary.get("status_301", 0)],
        ["HTTP 404", summary.get("status_404", 0)],
        ["HTTP 5xx", summary.get("status_5xx", 0)],
        ["Missing Title", summary.get("missing_title", 0)],
        ["Missing Description", summary.get("missing_description", 0)],
        ["Missing H1", summary.get("missing_h1", 0)],
        ["Missing Canonical", summary.get("missing_canonical", 0)],
        ["Non-Indexable", summary.get("non_indexable", 0)],
        ["Avg Word Count", summary.get("avg_word_count", 0)],
    ]
    ws_summary.append(["Metric", "Value"])
    _apply_header_style(ws_summary, 2)
    for label, val in summary_data:
        ws_summary.append([label, val])
    _auto_width(ws_summary)

    # Sheet 3: Errors
    ws_errors = wb.create_sheet("Errors")
    error_rows = [r for r in results if r.get("error")]
    _write_rows(ws_errors, ["URL", "Status Code", "Error"], [
        [r.get("url", ""), r.get("status_code", ""), r.get("error", "")]
        for r in error_rows
    ])
    _auto_width(ws_errors)

    filename = f"sitemap-export-{domain.replace('.', '-')}-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath


def generate_tags_xlsx(tags: list[dict], recommendations: list[dict],
                       page_coverage: list[dict], domain: str) -> str:
    """Generate tag discovery XLSX. Returns file path."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    wb = Workbook()

    # Sheet 1: Tags Found
    ws_tags = wb.active
    ws_tags.title = "Tags Found"
    tag_headers = [
        "Vendor", "Vendor Short", "Category", "Tag ID", "Pages Found On",
        "Page Count", "Load Method", "DOM Location", "Script URL",
        "Code Snippet", "Async/Defer", "Duplicate?", "Impact", "Impact Detail", "Notes",
    ]
    ws_tags.append(tag_headers)
    _apply_header_style(ws_tags, len(tag_headers))
    for t in tags:
        ws_tags.append([
            t.get("vendor", ""),
            t.get("vendor_short", ""),
            t.get("category", ""),
            t.get("tag_id", ""),
            t.get("pages_found_on", ""),
            t.get("page_count", 0),
            t.get("load_method", ""),
            t.get("dom_location", ""),
            t.get("script_url", ""),
            (t.get("code_snippet", "") or "")[:500],
            t.get("async_defer", ""),
            t.get("is_duplicate", "No"),
            t.get("impact", ""),
            t.get("impact_detail", ""),
            t.get("notes", ""),
        ])
    _auto_width(ws_tags)

    # Sheet 2: Recommendations
    ws_recs = wb.create_sheet("Recommendations")
    rec_headers = ["Priority", "Category", "Recommendation", "Affected Tags", "Expected Impact"]
    ws_recs.append(rec_headers)
    _apply_header_style(ws_recs, len(rec_headers))
    for r in recommendations:
        ws_recs.append([
            r.get("priority", ""),
            r.get("category", ""),
            r.get("recommendation", ""),
            r.get("affected_tags", ""),
            r.get("expected_impact", ""),
        ])
        # Color-code priority
        row_idx = ws_recs.max_row
        prio = r.get("priority", "").lower()
        if prio == "critical":
            ws_recs.cell(row=row_idx, column=1).fill = ERROR_FILL
        elif prio == "high":
            ws_recs.cell(row=row_idx, column=1).fill = WARNING_FILL
    _auto_width(ws_recs)

    # Sheet 3: Page Coverage
    ws_pages = wb.create_sheet("Page Coverage")
    page_headers = ["Page Type", "URL", "Status", "Tags Found", "GTM", "GA4", "Meta Pixel", "Other Tags"]
    ws_pages.append(page_headers)
    _apply_header_style(ws_pages, len(page_headers))
    for p in page_coverage:
        ws_pages.append([
            p.get("page_type", ""),
            p.get("url", ""),
            p.get("status", ""),
            p.get("tags_found", 0),
            p.get("gtm", "No"),
            p.get("ga4", "No"),
            p.get("meta_pixel", "No"),
            p.get("other_tags", ""),
        ])
    _auto_width(ws_pages)

    filename = f"tag-discovery-{domain.replace('.', '-')}-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath
